from openpyxl.reader.excel import load_workbook


class ExcelHelper:

    @staticmethod
    def get_test_data(file_path, sheet_name, test_name):
        wb = load_workbook(file_path)
        sheet = wb[sheet_name]
        header =[]
        #read data
        for cell in sheet[1]:
            header.append(cell.value)

        test_data ={}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = dict(zip(header, row))
            if row_data["Test Name"] == test_name:
                test_data = row_data
                break
        wb.close()
        return test_data


